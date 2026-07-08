"""
Servicio: generacion de plantillas Excel para carga masiva.
Usa openpyxl para crear archivos .xlsx con formato y cabeceras predefinidas.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


class TemplateGenerator:

    @staticmethod
    def generate_employee_template(dest_path):
        """
        Genera un archivo .xlsx con la plantilla para carga masiva de empleados.
        La primera fila contiene las cabeceras formateadas; las siguientes son
        filas de ejemplo que el usuario puede reemplazar con datos reales.
        @param dest_path ruta de destino donde se guardara el archivo .xlsx
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empleados"

        cabeceras = ["cedula", "nombre", "telefono", "correo"]
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, nombre_col in enumerate(cabeceras, start=1):
            celda = ws.cell(row=1, column=col_idx, value=nombre_col)
            celda.fill = header_fill
            celda.font = header_font
            celda.alignment = header_alignment

        # Fila de ejemplo comentada para guiar al usuario
        ejemplo = ["12345678", "Juan Perez", "04141234567", "juan@ejemplo.com"]
        for col_idx, valor in enumerate(ejemplo, start=1):
            celda = ws.cell(row=2, column=col_idx, value=valor)
            celda.font = Font(italic=True, color="808080")

        # Ancho de columnas
        anchos = {"A": 16, "B": 30, "C": 18, "D": 32}
        for col_letra, ancho in anchos.items():
            ws.column_dimensions[col_letra].width = ancho

        ws.row_dimensions[1].height = 20

        wb.save(dest_path)
